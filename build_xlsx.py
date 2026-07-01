import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC='/mnt/user-data/outputs/muttcliffe_sheet/'
OUT='/mnt/user-data/outputs/MUTTCLIFFE_DepthChart_Data.xlsx'
DATA_TABS=['Teams','Units','Players','Callouts']   # order in the workbook

HEAD_FILL=PatternFill('solid', fgColor='11202E')
HEAD_FONT=Font(name='Arial', bold=True, color='FFFFFF', size=10)
CELL_FONT=Font(name='Arial', size=10)
KEY_FILL =PatternFill('solid', fgColor='EAF0F6')
thin=Side(style='thin', color='D5DCE4')
BORDER=Border(bottom=thin, right=thin)

wb=Workbook()
wb.remove(wb.active)

# ---- READ ME tab ----
ws=wb.create_sheet('_READ ME')
lines=[
 ('MUTTCLIFFE — Depth Chart data workbook',True),
 ('',False),
 ('This workbook is the single source of truth for the hosted depth-chart app. Four data tabs, keyed so every',False),
 ('team and every week is a set of rows. Edit here; the app reads it live (no code changes, no redeploys).',False),
 ('',False),
 ('TABS',True),
 ('  Teams     — one row per team+week: header meta, verdict, ratio meter, TL;DR headline/trajectory.',False),
 ('  Units     — one row per team+week+unit (10 units/team): tier / conf / depth / flag / note.',False),
 ('  Players   — one row per player at a slot (only teams with a full roster; Ottawa is the reference).',False),
 ('  Callouts  — the TL;DR "dragging the rating" (drag) and "watch & returns" (watch) lists.',False),
 ('  Slots     — reference only: valid slot IDs + where each sits on the field. The app ignores this tab.',False),
 ('',False),
 ('GOLDEN RULES',True),
 ('  1. One row = one record. Never merge cells. Keep the header row (row 1) exactly as-is.',False),
 ('  2. Weekly update = a DELTA. Add a new week by copying a team\'s current rows and bumping the "week"',False),
 ('     column; then change only the cells that moved. The app shows the latest week automatically.',False),
 ('  3. Standing absences stay as they are — don\'t re-enter an unchanged roster each week.',False),
 ('  4. A team with NO Players rows renders at unit level (tinted blocks). Add Players rows to light up chips.',False),
 ('  5. Values can be blank; leave the cell empty (don\'t type "—" unless you want it shown).',False),
 ('',False),
 ('GO LIVE (one time)',True),
 ('  1. Share this sheet: File ▸ Share ▸ General access ▸ "Anyone with the link" = Viewer.',False),
 ('  2. Copy the sheet ID from the URL (the long string between /d/ and /edit).',False),
 ('  3. In muttcliffe_depthchart.html set  SHEET_ID = "<that id>"  and  DATA_SOURCE = "sheet".',False),
 ('  4. Reload the app — the source label should read "google sheet · N teams".',False),
 ('',False),
 ('The team chats emit paste-ready rows in exactly this column order (see the Weekly Export appendix).',False),
]
for i,(txt,bold) in enumerate(lines,1):
    c=ws.cell(row=i,column=1,value=txt); c.font=Font(name='Arial', bold=bold, size=12 if (bold and i==1) else 10)
ws.column_dimensions['A'].width=112
ws.sheet_view.showGridLines=False

# ---- data tabs ----
NOTEISH={'note','headline','traj_note','flag'}
WIDES={'note':46,'headline':70,'traj_note':52,'flag':22,'name':13,'full':18,'pos':11,'persist':22,
       'verdict_persist':26,'verdict_mvar':26,'verdict_ratio':26,'label':14,'venue':22,'oppFull':22}
for tab in DATA_TABS:
    df=pd.read_csv(SRC+tab+'.csv', dtype=str, keep_default_na=False)
    ws=wb.create_sheet(tab)
    ws.append(list(df.columns))
    for _,row in df.iterrows():
        ws.append(['' if (v is None or str(v)=='nan') else str(v) for v in row.tolist()])
    # style header
    for ci,col in enumerate(df.columns,1):
        h=ws.cell(row=1,column=ci); h.fill=HEAD_FILL; h.font=HEAD_FONT; h.alignment=Alignment(vertical='center')
        w=WIDES.get(col, 12 if col in ('team','week','side','slot','depth','num','nat','kind','order','key','tier','conf') else 14)
        ws.column_dimensions[get_column_letter(ci)].width=w
    # body font + key-column tint + wrap for noteish
    for r in range(2, ws.max_row+1):
        for ci,col in enumerate(df.columns,1):
            cell=ws.cell(row=r,column=ci); cell.font=CELL_FONT
            if col in ('team','week'): cell.fill=KEY_FILL
            if col in NOTEISH: cell.alignment=Alignment(wrap_text=True, vertical='top')
            else: cell.alignment=Alignment(vertical='top')
    ws.freeze_panes='A2'
    ws.row_dimensions[1].height=18
    ws.sheet_view.showGridLines=True

# ---- Slots reference tab ----
ws=wb.create_sheet('Slots')
ws.append(['slot','side','label','row','col','note'])
SLOTS=[
 ('WR_B','offense','WR ⟵',0,0,'wide receiver, boundary'),('SB_1','offense','SB',0,1,'slotback'),
 ('SB_2','offense','SB',0,2,'slotback'),('SB_3','offense','SB',0,3,'slotback'),('WR_F','offense','⟶ WR',0,4,'wide receiver, field'),
 ('LT','offense','LT',1,0,'left tackle'),('LG','offense','LG',1,1,'left guard'),('C','offense','C',1,2,'centre'),
 ('RG','offense','RG',1,3,'right guard'),('RT','offense','RT',1,4,'right tackle'),
 ('QB','offense','QB',2,0,'quarterback'),('RB','offense','RB',2,1,'running back'),('FB','offense','FB',2,2,'fullback'),
 ('CB_B','defense','CB ⟵',0,0,'corner, boundary'),('HB_B','defense','HB',0,1,'halfback'),('S','defense','S',0,2,'safety'),
 ('HB_F','defense','HB',0,3,'halfback'),('CB_F','defense','⟶ CB',0,4,'corner, field'),
 ('WLB','defense','WLB',1,0,'weak-side LB'),('MLB','defense','MLB',1,1,'middle LB'),('SLB','defense','SLB',1,2,'strong-side LB'),
 ('DE_B','defense','DE ⟵',2,0,'edge, boundary'),('DT_1','defense','DT',2,1,'tackle'),('DT_2','defense','DT',2,2,'tackle'),('DE_F','defense','⟶ DE',2,3,'edge, field'),
 ('KP','st','K / P',None,None,'kicker / punter'),('LS','st','LS',None,None,'long snapper'),('RET','st','RET',None,None,'returner'),
 ('(use side=il)','il','—',None,None,'deep IR not dressed; leave slot blank, side=il'),
]
for r in SLOTS: ws.append(['' if v is None else v for v in r])
for ci,col in enumerate(['slot','side','label','row','col','note'],1):
    h=ws.cell(row=1,column=ci); h.fill=HEAD_FILL; h.font=HEAD_FONT
for r in range(2, ws.max_row+1):
    for ci in range(1,7): ws.cell(row=r,column=ci).font=CELL_FONT
for col,w in zip('ABCDEF',[14,10,10,7,7,34]): ws.column_dimensions[col].width=w
ws.freeze_panes='A2'

# ---- Inbox tab (combined-paste weekly workflow; the app ignores this tab) ----
ws=wb.create_sheet('Inbox')
head=['tab','team','week','←  paste combined export rows below; fields follow each tab\u2019s column order  →']
ws.append(head)
for ci in range(1, 34):                      # 33 cols = widest record (Players 32) + leading tab column
    ws.cell(row=1, column=ci).fill=HEAD_FILL
for ci,txt in enumerate(head,1):
    c=ws.cell(row=1,column=ci); c.font=HEAD_FONT; c.alignment=Alignment(vertical='center')
for col,w in zip('ABC',[12,10,8]): ws.column_dimensions[col].width=w
ws.column_dimensions['D'].width=70
notes=[
 ['','','','Each team chat emits ONE block; every row starts with its tab name (Teams / Units / Players / Callouts).'],
 ['','','','Paste all blocks here, then run  MUTTCLIFFE ▸ Process inbox  — rows fan out to the right tabs and this clears.'],
 ['','','','Re-pasting a corrected (team, week) block replaces the old rows automatically (idempotent).'],
]
for r in notes:
    ws.append(r)
    ws.cell(row=ws.max_row, column=4).font=Font(name='Arial', italic=True, size=9, color='6B7A8D')
ws.freeze_panes='A2'
ws.row_dimensions[1].height=18
ws.sheet_view.showGridLines=True

wb.save(OUT)
print('saved', OUT)
print('tabs:', wb.sheetnames)
